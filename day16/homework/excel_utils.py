"""
-------------------------------------------------
   File Name:excel_utils
   Author:Lee
   date: 2021/6/16-10:40
-------------------------------------------------
"""
import openpyxl


class CaseData(object):
    def __init__(self, dict_case):
        for i in dict_case.items():
            setattr(self, i[0], i[1])  # 利用反射添加属性


class ReadExcel(object):
    @classmethod
    def read_data_all(cls, file_name, sheet_name):  # 读取所有的内容
        wb = openpyxl.load_workbook(file_name)  # 读取工作簿
        sh = wb[sheet_name]  # 获取sheet页
        rows = list(sh.rows)  # 获取所有表格
        # print(rows)
        allCase = []
        # 读取表头数据
        titles = []
        for cell in rows[0]:  # rows[0]代表第一行数据
            titles.append(cell.value)  # 将第一行cell的值添加到titles中
        # print(titles)  # ['case_id', 'case_title', 'interface', 'url', 'case_data', 'expect', 'result']

        # 遍历其他行数据，和表头打包，转换为字典，放到列表
        for row in rows[1:]:  # row代表除了表头的每一行数据
            data = []  # [1,登录成功，login....]
            for v in row:
                data.append(v.value)
            case_zip = dict(zip(titles, data))  # 将数据进行打包
            case_data = CaseData(case_zip)  # 调用构造方法，将数据转换为对象
            allCase.append(case_data)

        return allCase

    @classmethod
    def read_data_pl(cls, file_name, sheet_name, begin_row, end_row):
        wb = openpyxl.load_workbook(file_name)
        sh = wb[sheet_name]
        rows = list(sh.rows)
        allCase = []

        # 读取表头数据
        titles = []
        for cell in rows[0]:
            titles.append(cell.value)

        # 遍历其他行数据，和表头打包，转换为字典，存放到列表中
        for row in rows[begin_row:end_row + 1]:
            data = []
            for v in row:
                data.append(v.value)
            case_zip = dict(zip(titles, data))
            cd = CaseData(case_zip)
            allCase.append(cd)
        return allCase

    @classmethod
    def write_data(cls, file_name, sheet_name, row, column, value):
        """
        :param file_name: 需要读写文件
        :param sheet_name: 需要读写的sheet页
        :param row: 需要被写入的行
        :param column: 需要被写入列
        :param value: 需要写入的内容
        """
        wb = openpyxl.load_workbook(file_name)
        sh = wb[sheet_name]
        sh.cell(row=row + 1, column=column, value=value)
        wb.save(file_name)


if __name__ == '__main__':
    # 获取所有内容
    cases = ReadExcel.read_data_all(r'./cases.xlsx', 'Login')
    print(cases)  # 存放了数据对象的列表
    for case in cases:
        print(case.case_id, case.url, case.case_data, case.expect)

    # 获取部分内容
    cases = ReadExcel.read_data_pl(r'./cases.xlsx', 'Login', 2, 2)
    print(cases)
    for case in cases:
        print(case.case_id, case.url, case.case_data, case.expect)

    cases = ReadExcel.read_data_pl(r'./cases.xlsx', 'Login', 2, 3)
    print(cases)
    for case in cases:
        print(case.case_id, case.url, case.case_data, case.expect)

    # 写入数据
    ReadExcel.write_data(r'./cases.xlsx', 'Login', 2, 7, 'FAILED')
