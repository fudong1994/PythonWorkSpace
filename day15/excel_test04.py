"""
-------------------------------------------------
   File Name:excel_test04
   Author:Lee
   date: 2021/6/15-16:51
-------------------------------------------------
"""
"""
python操作excel
"""
import openpyxl  # 用来操作Excel文件

workbook = openpyxl.load_workbook(r'./cases.xlsx')  # 用来读取excel并且看作一个对象
sheet = workbook['Login']  # 获取sheet页

# 获取cell中的值 (excel文件读出的数据类型除了 int/float类型外，全部都是str)
res = sheet.cell(row=2, column=5).value
print(res)  # {'username':'xiaohua','password':'a123456'}
print(type(res))  # <class 'str'>
print(list(sheet.rows))  # 获取所有的cell

# 写数据
sheet.cell(row=2, column=7).value = '19班的第一次excel读写'
workbook.save(r'./cases.xlsx')
